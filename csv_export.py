import os
import re
import shutil
import calendar
from datetime import datetime
from pathlib import Path

import chardet
import pandas as pd
from openpyxl import Workbook

# Constantes
INPUT_DIR = Path("input_csvs")
PROCESSED_DIR = Path("processed")
OUTPUT_DIR = Path("output")

MONTHS_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

DATE_ALIASES = {"date", "fecha", "day", "time", "datetime"}
EARN_ALIASES = {"earnings", "earning", "revenue", "profit", "amount", "amount (btc)", "amount(btc)", "btc"}

# --------------------
# Helpers
# --------------------
def detect_encoding(file_path: Path) -> str:
    with file_path.open("rb") as f:
        raw = f.read(200000)
    return chardet.detect(raw)["encoding"] or "utf-8"

def norm_header(s: str) -> str:
    return " ".join(s.strip().lower().split())

def choose_column(df: pd.DataFrame, aliases: set) -> str:
    norm_map = {c: norm_header(c) for c in df.columns}
    for c, n in norm_map.items():
        if n in aliases or any(a in n for a in aliases):
            return c
    return None

def read_csv_smart(file_path: Path) -> pd.DataFrame:
    enc = detect_encoding(file_path)
    for sep in [",", ";", "\t", "|"]:
        try:
            df = pd.read_csv(file_path, encoding=enc, sep=sep, engine="python", dtype=str, keep_default_na=False)
            if df.shape[1] > 1:
                return df
        except pd.errors.ParserError:
            continue
    return pd.read_csv(file_path, encoding=enc, dtype=str, keep_default_na=False)

def to_float_amount(x: str) -> float:
    x = x.strip().replace(" ", "")
    x = x.replace(",", ".") if "," in x else x
    try:
        return float(x)
    except ValueError:
        return 0.0

def safe_filename(text: str) -> str:
    return re.sub(r'[\/:*?"<>|]', "_", text.strip())[:60]

def sanitize_sheet_name(name: str) -> str:
    return re.sub(r"[:\\/?*\[\]]", "_", name).strip()[:31] or "Subcuenta"

def apply_number_formats(ws, header_to_col):
    formats = {
        "Amount (BTC)": "0.00000000",
        "BTC Price": "0.00",
        "USD": "0.00",
        "Date": "dd/mm/yyyy"
    }
    for col, fmt in formats.items():
        for r in range(2, ws.max_row + 1):
            ws.cell(r, header_to_col[col]).number_format = fmt

# --------------------
# Main
# --------------------
def main():
    year, month = get_year_month()
    BTC_USD_RATE = get_btc_usd_rate()
    subcuenta = input("Ingrese subcuenta: ").strip()
    
    if not subcuenta or not os.listdir(INPUT_DIR):
        print("Subcuenta obligatoria o no hay CSVs en input_csvs.")
        return

    files = [f for f in INPUT_DIR.glob("*.csv")]

    all_out = []
    processed_files = []

    for file_path in files:
        df = read_csv_smart(file_path)
        date_col = choose_column(df, DATE_ALIASES)
        earn_col = choose_column(df, EARN_ALIASES)

        if not date_col or not earn_col:
            continue

        mask = (pd.to_datetime(df[date_col], errors="coerce", dayfirst=True).dt.year == year) & \
               (pd.to_datetime(df[date_col], errors="coerce", dayfirst=True).dt.month == month)

        if not mask.any():
            continue

        earn = df.loc[mask, earn_col].apply(to_float_amount)
        out = pd.DataFrame({
            "Subcuenta": subcuenta,
            "Description": "Miner Revenue",
            "Date": df.loc[mask, date_col],
            "Amount (BTC)": earn,
            "BTC Price": BTC_USD_RATE,
            "USD": earn * BTC_USD_RATE
        })

        all_out.append(out)
        processed_files.append(file_path)

    if not all_out:
        print("No se encontraron datos del mes solicitado.")
        return

    save_to_excel(all_out, year, month, subcuenta)
    move_processed_files(processed_files, month, subcuenta)
        move_processed_files(processed_files, month, subcuenta)

def get_year_month():
    while True:
        try:
            year = int(input("Ingrese año (ej: 2026): ").strip())
            month = int(input("Ingrese mes (1-12): ").strip())
            if 1 <= month <= 12:
                return year, month
        except ValueError:
            print("Valor inválido.")

def get_btc_usd_rate():
    while True:
        try:
            rate_input = input("Ingrese BTC → USD rate: ").strip()
            return float(rate_input.replace(",", "."))
        except ValueError:
            print("Valor inválido.")

def save_to_excel(all_out, year, month, subcuenta):
    out_all = pd.concat(all_out, ignore_index=True).sort_values("Date")
    
    mes_nombre = f"{month:02d}_{MONTHS_ES[month-1]}"
    subcuenta_dir = OUTPUT_DIR / safe_filename(subcuenta)
    os.makedirs(subcuenta_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = subcuenta_dir / f"Transacciones_Minado_{mes_nombre}_{year}_{ts}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_name(subcuenta)

    for c, col in enumerate(out_all.columns, start=1):
        ws.cell(1, c, col)

    for r, row in enumerate(out_all.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)

    add_totals(ws, out_all)
    apply_number_formats(ws, {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)})
    validate_days(ws, year, month, out_all)

    wb.save(out_path)
    print(f"\n✅ Excel creado: {out_path}")

def add_totals(ws, out_all):
    last_data_row = ws.max_row
    total_row = last_data_row + 1
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    ws.cell(total_row, header_to_col["Date"], "Total")
    ws.cell(total_row, header_to_col["Amount (BTC)"],
        f"=SUBTOTAL(109,{ws.cell(2, header_to_col['Amount (BTC)']).coordinate}:{ws.cell(last_data_row, header_to_col['Amount (BTC)']).coordinate})")
    ws.cell(total_row, header_to_col["USD"],
        f"=SUBTOTAL(109,{ws.cell(2, header_to_col['USD']).coordinate}:{ws.cell(last_data_row, header_to_col['USD']).coordinate})")

def validate_days(ws, year, month, out_all):
    last_day = calendar.monthrange(year, month)[1]
    present_days = set(pd.to_datetime(out_all["Date"]).dt.day.unique())
    missing = [d for d in range(1, last_day + 1) if d not in present_days]

    ws2 = ws.parent.create_sheet("Control")
    ws2["A1"] = "Periodo"
    ws2["B1"] = f"{MONTHS_ES[month-1]} {year}"
    ws2["A3"] = "Días faltantes"
    ws2["B3"] = ", ".join(f"{d:02d}" for d in missing) if missing else "Mes completo"

def move_processed_files(processed_files, month, subcuenta):
    proc_dir = PROCESSED_DIR / f"{month:02d}_{MONTHS_ES[month-1]}" / safe_filename(subcuenta)
    os.makedirs(proc_dir, exist_ok=True)

    for f in processed_files:
        shutil.move(str(f), str(proc_dir / f.name))

if __name__ == "__main__":
    main()